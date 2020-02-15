# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 12:34:05 2020

@author: satye
"""

import openpyxl
import os
os.chdir('C:\\users\\satye.MSI\\Documents\\GitHub\\Resume is LIT')

workbook = openpyxl.load_workbook('record.xlsx')
print(type(workbook))
#sheet = workbook.get_sheet_by_name('Sheet1') #if we know sheet names
sheet = workbook['Sheet1']
#workbook.get_sheet_names()
print(workbook.sheetnames) #display sheetnames
cell = sheet['A2'] #returns a cell object
print(cell.value) #returns integer value
print(str(cell.value)) #returns string

print(sheet.cell(row=1,column=2))

for i in range(1,8):
    print(i, sheet.cell(row=i,column=2).value)