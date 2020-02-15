# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 20:11:04 2020

@author: satye
"""

import openpyxl,os
wb = openpyxl.Workbook() #create a new workbook object
print(wb.sheetnames) #defualt created sheet in workbook
sheet = wb['Sheet']
print(sheet)
print(sheet['A1'].value)
print(sheet['A1'].value == None)

sheet['A1'] = 42
sheet['A2'] = 'Hello' # these are stored in temp memory of python code
os.chdir('C:\\users\\satye.MSI\\Documents\\GitHub\\Resume is LIT')
wb.save('example.xlsx')
sheet2 = wb.create_sheet() #adding sheet into the xlsx file
print(wb.sheetnames)
print(sheet2.title)
sheet2.title = 'New Sheet Name'
print(wb.sheetnames)
wb.save('example2.xlsx') #dont overwrite
wb.create_sheet(index=0,title = 'My other sheet') #create a sheet at a particular order
wb.save('example3.xlsx')